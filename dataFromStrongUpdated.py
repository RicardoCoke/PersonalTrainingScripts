#Imports
from warnings import catch_warnings
from numpy import number
import pandas as pd
import datetime
import csv
from pathlib import Path
import re
import xlsxwriter
from xlsxwriter.workbook import Workbook
import numpy as np
import sys

##Functions-------------
def updateLine(exerciseName, chosenLine):

    #need to append to corresponding exercise line array    
    if(exerciseName.__contains__("Bench Press") or exerciseName.__contains__("Supino")):
        bpLine.append(chosenLine + 1)
    
    elif(exerciseName.__contains__("Squat") or exerciseName == "Leg Press" or exerciseName == "Agachamento"):
        squatLine.append(chosenLine + 1)

    elif(exerciseName == "Rowing (Machine)" or exerciseName == "Bent Over Row (Barbell)" or exerciseName == "Remada"):
        rowLine.append(chosenLine + 1)

    elif(exerciseName.__contains__("Pull Up") or exerciseName == "Chin Up"):
        pullUpLine.append(chosenLine + 1)

    elif(exerciseName == "Shoulder Press (Machine)" or exerciseName =="Seated Overhead Press (Dumbbell)"):
        spLine.append(chosenLine + 1)

    elif(exerciseName.__contains__("Deadlift") or exerciseName == "Peso morto stiff"):
        stiffLine.append(chosenLine + 1)

    elif(exerciseName == "Hip Thrust (Barbell)"):
        hpLine.append(chosenLine + 1)

    elif(exerciseName == "Lunge"):
        lungeLine.append(chosenLine + 1)

#Transforms similiar movement in valid exercise Name
def exerciseNameVerifier(exerciseName):

    if(exerciseName.__contains__("Bench Press") or exerciseName.__contains__("Supino")):

         exerciseName = "Bench Press (Barbell)"
         return exerciseName
    
    elif(exerciseName.__contains__("Squat") or exerciseName.__contains__("Leg Press") or exerciseName.__contains__("Agachamento")):

         exerciseName = "Squat (Barbell)"
         return exerciseName

    elif(exerciseName.__contains__("Rowing") or exerciseName.__contains__("Bent Over Row") or exerciseName.__contains__("Remada")):

        exerciseName = "Rowing (Machine)"
        return exerciseName
        
    elif(exerciseName.__contains__("Pull Up") or exerciseName.__contains__("Chin Up") or exerciseName.__contains__("Pulldown")):
        exerciseName = "Pull Up"
        return exerciseName

    elif(exerciseName.__contains__("Shoulder Press") or exerciseName.__contains__("Seated Overhead Press")):
        exerciseName = "Shoulder Press (Machine)" 
        return exerciseName

    elif(exerciseName.__contains__("Deadlift") or exerciseName.__contains__("Peso morto")):
        exerciseName = "Stiff Leg Deadlift (Barbell)"
        return exerciseName

    elif(exerciseName.__contains__("Hip Thrust") or exerciseName.__contains__("Glute Kickback")):
        exerciseName = "Hip Thrust (Barbell)"
        return exerciseName

    elif(exerciseName.__contains__("Lunge")):
        exerciseName = "Lunge"
        return exerciseName
         
    else:
        exerciseName = "invalid"
        return exerciseName
    
def strongConverter(readArr,i,ii):

    if(readArr[i][ii] == "Exercise"):
        readArr[i][ii] = "Exercise Name"

    if(readArr[i][ii] == "Weight (kgs)"):
        readArr[i][ii] = "Weight"

    if(readArr[i][ii] == "Category"):
        readArr[i][ii] = "Set Order"   

def fitNotesConverter(fileName):
    # Row Array
    readArr = []
    # CSV reader
    with open(fileName, newline='') as file:
        reader = csv.reader(file, delimiter=',', quotechar='|')
        for row in reader:
            #Store all the rows in readArr to manipulate further
            readArr.append(row)

    # Verifies if Data was already Converted Before (Row 1, Column 7 should be "wasConverted")
    if (readArr[1][7] == "wasConverted"):
        print("Data was already converted before.")
        return

    # Iterate over the readArr , change first row (HEADERS) to StrongApp Headers, add Set Order after Exercise

    # Convert Fitnotes CSV into same Headers as StrongApp
    # i represents the first row of the CSV (first line)
    for i in range(len(readArr)):
        for ii in range(len(readArr[i])):
            strongConverter(readArr,i,ii)
            
    # Replace converted CSV Category Header information with the correct Set Order Expected
    setOrder = 1
    row = 1
    while(row < len(readArr)):

        # First row case
        if(row == 1):
            readArr[row][2] = setOrder
            readArr[row][6] = "" # RPE
            readArr[row][7] = "wasConverted"
            previousExercise = readArr[row][1]
            row +=1
            continue # Re-start the for cycle

        # Rest of the rows
        if(readArr[row][1] == previousExercise):
            setOrder +=1
            readArr[row][2] = setOrder
            readArr[row][6] = "" #RPE
            row +=1

        else:
            setOrder = 1
            readArr[row][2] = setOrder
            readArr[row][6] = "" #RPE
            previousExercise = readArr[row][1]
            row +=1
    
    #Write the modified Info in the CSV file
    with open(fileName, 'w', newline='') as csvfile: 
        # creating a csv writer object 
        csvwriter = csv.writer(csvfile)   

        # writing the data rows 
        csvwriter.writerows(readArr)

def check_data_validity(fileName):
    with open(fileName, newline = "") as csvfile:
        try:
            dialect = csv.Sniffer().sniff(csvfile.read(1024), delimiters = ";")
            print("Delimiter is ;")

            #Find and replace ; to ,
            text = open(fileName, "r")
            text = ''.join([i for i in text]) \
                .replace(";", ",")

            x = open(fileName,"w")
            x.writelines(text)
            x.close()

        except:
            print("Delimiter is already a ,")

def removeHeadersStrong():
    # Verify if data was digested before
    for i in range(0, len(data[0])):
        if (data[0][i] == "wasConverted"):
            print("Data was converted before.")
            return

    #Delete unecessary Headers
    del data[0][1] 
    del data[0][4]

    for i in range(1,len(data)):

        # Remove hours from Date
        prevDate = data[i][0]
        newOnlyDate = re.split('\s.*', prevDate)[0]
        data[i][0] = newOnlyDate

        #Remove unecessary rows (Workout Name, Weight Unit)
        del data[i][1] #Workout name
        del data[i][4] #Weight Unit

    # Add Mark that data was digested before (wasConverted)
    data[0][9] = "wasConverted"

    #Write the modified Info in the CSV file
    with open(users[userInput], 'w', newline='') as csvfile: 
        # creating a csv writer object 
        csvwriter = csv.writer(csvfile, quotechar = '?')   

        # writing the data rows 
        csvwriter.writerows(data)
        return

def fillDataObject(data):
    with open(users[userInput], newline='') as file:
        reader = csv.reader(file, delimiter=',', quotechar='|')
        for row in reader:
            #Store all the rows in readArr to manipulate further
            data.append(row)

############################################################
##########################################################

# Users hashtable
users = {"josemiguel": r"FitNotes_Export josemiguel.csv","ruimoreira": r"strong_ruimoreira.csv", 
"inesbarros": "FitNotes_Export inesbarros.csv", "catarinaferreira": r"strong_catarinaferreira.csv",
"teste": r"strong_teste.csv", "rui": r"strong_rui.csv", "rui_teste": "strong_rui_teste.csv",
"joaodias": r"strong_joaodias.csv", "carlos": r"strong_carlos.csv", "samir": r"strong_samir.csv"}

# Exercise List
exerciseList = ["Bench Press (Barbell)","Squat (Barbell)","Rowing (Machine)","Pull Up",
"Shoulder Press (Machine)","Stiff Leg Deadlift (Barbell)","Hip Thrust (Barbell)", "Lunge"]

#Ask for user input manually if not introduced in command line
if(len(sys.argv) < 2):
    userInput = str(input("Enter the user you want to update.\n")).lower()
else:
    userInput = sys.argv[1]

#Insert File Name Manually
#userInput = "rui"

#Checks if the delimiter is a comma (,), if it isnt, converts ; to ,
check_data_validity(users[userInput])

# Fills the Data array with the info from CSV
data = []
fillDataObject(data)

# PRE-DATA DIGEST TO TRANSFORM FITNOTES MODEL INTO STRONG CSV
if('FitNotes_Export' in users[userInput]):
    fitNotesConverter(users[userInput])
    print("Pre-processing done")

# PRE-DATA DIGEST TO REMOVE HOURS INCASE OF STRONG EXPORT CSV
else:
    removeHeadersStrong()

# Re-Update Data object with CSV modifications
fillDataObject(data)

####################################################################

# Dates for validation
today = datetime.date.today()
todayObject = datetime.datetime(today.year, today.month, today.day)
lastWeek = today - datetime.timedelta(days=7) 
lastMonthDay = today - datetime.timedelta(days=28) 
lastMonthDayObject = datetime.datetime(lastMonthDay.year, lastMonthDay.month, lastMonthDay.day)

#Data,Nome do treino,Duração,Nome do exercício,Ordem da série,Peso,Reps,Distância,Segundos,Notas,Notas do treino,RPE
#Dates array
rowIndex = []
for rowi in range(1,len(data)):

    #Regex to isolate date from other stuff
    onlyDate = data[rowi][0]
    if(onlyDate == "Date"):
        break


    #Gets the day and year from the DataFrame object
    day = int(re.split('-',onlyDate)[2])
    month = int(re.split('-',onlyDate)[1])
    year = int(re.split('-',onlyDate)[0])
    actualDateTimeObject = datetime.datetime(year, month, day) #row date datetime object

    if(actualDateTimeObject >= lastMonthDayObject and actualDateTimeObject <= todayObject):
        rowIndex.append(rowi) #this value doesnt take into account first line(headers)


#Algorithm to store exercise performance
#Create the xlsx sheet in case doesnt exist

#Writes data to user CSV performance review file
filename = userInput +'_perfReview_' + str(today.day) + '_' + str(today.month) + '_' + str(today.year) + '.xlsx'
workbook = xlsxwriter.Workbook(filename)
worksheet = workbook.add_worksheet("Dados_Treino")

#Column and Line initialization
chosenArr = []
chosenLine = 1
bpLine = [3]
squatLine = [3]
rowLine = [3]
pullUpLine = [3]
spLine = [3]
stiffLine = [3]
hpLine = [3]
lungeLine = [3]
column = 1
line = 1

#Header format
headFormat = workbook.add_format()
headFormat.set_bold()
headFormat.set_bg_color("#366092")
headFormat.set_font_color("white")

#Sub-Header format 
subHeadFormat = workbook.add_format()
subHeadFormat.set_bg_color("#95B3D7")
subHeadFormat.set_font_color("white")

#Relative Header format (with width adjust)
relHeadWidthFormat = workbook.add_format()
relHeadWidthFormat.set_bold()
relHeadWidthFormat.set_bg_color("#963634")
relHeadWidthFormat.set_font_color("white")

#Relative Header format (without width adjust)
relHeadFormat = workbook.add_format()
relHeadFormat.set_bold()
relHeadFormat.set_bg_color("#963634")
relHeadFormat.set_font_color("white")

#Sub-Relative Perf header format 
subRelHeadFormat = workbook.add_format()
subRelHeadFormat.set_bg_color("#DA9694")
subRelHeadFormat.set_font_color("white")

#Exercise Headers Function
def writeHeaders():
    # Headers
    headerArr = np.arange(round(len(exerciseList) * 8.57))
    headIndex = 0

    for i in range(0, exerciseList.__len__()):
        worksheet.write(line + 1, headerArr[headIndex + 1], "Set", subHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 2], "Weight", subHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 3], "Reps", subHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 4], "RPE", subHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 5], "Data", subHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 6], "&Weight", subRelHeadFormat)
        worksheet.write(line + 1, headerArr[headIndex + 7], "&Reps", subRelHeadFormat)
        headIndex += 8

    # Exercise Names and Relative Performance
    i_header = 1
    for exercise in exerciseList:
        worksheet.write(line, i_header, exercise, headFormat)
        for iii in range(i_header+1, i_header+5):
            worksheet.write(line, iii, "", headFormat)
        worksheet.write(line, i_header+5, "Relative Performance", relHeadWidthFormat)
        worksheet.write(line, i_header+6, "", relHeadFormat)
        i_header += 8

# Writes the Headers (Exercise Name, Sets, Reps, etc)
writeHeaders()

#Get Initial exercise name and Date
exerciseName = exerciseNameVerifier(data[rowIndex[0]][1])
date = data[rowIndex[0]][0]

#Previous Date and Exercise
#TODO ERROR WHEN THE FIRST EXERCISE INSIDE OF DATE RANGE IS INVALID ONE
previousDate = date
previousExercise = exerciseName

#Store the exercises performance individually in according array
for i in range(len(rowIndex)):

    row = rowIndex[i]

    #Protection to avoid having different exerciseName even though same exercise with different nomenclature
    exerciseName = exerciseNameVerifier(data[row][1])

    if(exerciseName == "invalid"):
        #Need to make for cycle restart with i iterated
        continue

    date = data[row][0]
    serie = int(data[row][2])

    #Protection against invalid Data input from users
    if(data[row][3] == ''):
        data[row][3] = 1
        
    weight = float(data[row][3])

    #Protection to Pull-Up Weight being zero
    if(exerciseName == "Pull Up" and weight == 0):
        weight = 1.0

    try:
        reps = int(data[row][4])
    except Exception as e:
        print(e)
        print("Error occured in Date: {}; Reps input was: {}.".format(date, data[row][4]))
        try:
            reps = int(data[row][5])
        except Exception as e:
                 print(e)
                 print("Tried catching reps in the next column, error still occured.")

    rpe = data[row][5]
   

    #Exercise Columns (#TODO Refactor this into a function)
    bpCols = [1, 2, 3, 4, 5]
    squatCols = [9,10,11,12,13]
    rowCols = [17,18,19,20,21]
    pullUpCols = [25,26,27,28,29]
    spCols = [33,34,35,36,37]
    stiffCols = [41,42,43,44,45]
    hpCols = [49,50,51,52,53]
    lungeCols = [57,58,59,60,61]

    #Writes two empty Line to divide different date workout data
    if(previousDate != date or previousExercise != exerciseName):
        emptyLine = chosenLine+1
        worksheet.write(emptyLine, chosenArr[0], "")
        worksheet.write(emptyLine+1, chosenArr[0], "")
        updateLine(previousExercise, emptyLine+1)

    #Exercise present in actual line
    if(exerciseName.__contains__("Bench Press") or exerciseName.__contains__("Supino")):
        chosenArr = bpCols
        chosenLine = bpLine[-1] #Last Recorded Line for Exercise
    
    elif(exerciseName.__contains__("Squat") or exerciseName == "Leg Press"):
        chosenArr = squatCols
        chosenLine = squatLine[-1]

    elif(exerciseName == "Rowing (Machine)" or exerciseName == "Bent Over Row (Barbell)"):
        chosenArr = rowCols
        chosenLine = rowLine[-1]
        
    elif(exerciseName == "Pull Up" or exerciseName == "Chin Up"):
        chosenArr = pullUpCols
        chosenLine = pullUpLine[-1]

    elif(exerciseName == "Shoulder Press (Machine)"):
        chosenArr = spCols
        chosenLine = spLine[-1]

    elif(exerciseName.__contains__("Deadlift")):
        chosenArr = stiffCols   
        chosenLine = stiffLine[-1] 

    elif(exerciseName == "Hip Thrust (Barbell)"):
        chosenArr = hpCols   
        chosenLine = hpLine[-1]
    
    elif(exerciseName.__contains__("Lunge")):
        chosenArr = lungeCols   
        chosenLine = lungeLine[-1]


    #Add the series, weight, reps and RPE to the xlsx
    worksheet.write(chosenLine, chosenArr[0], serie)
    worksheet.write(chosenLine, chosenArr[1], weight)
    worksheet.write(chosenLine, chosenArr[2], reps)
    worksheet.write(chosenLine, chosenArr[3], rpe)
    worksheet.write(chosenLine, chosenArr[4], date)

    previousDate = date #Updates previousDate
    previousExercise = exerciseName
    updateLine(exerciseName, chosenLine) #Increments line after adding data


workbook.close()

######################### PART 2 - Volume calculation and cell coloring ###############################
#######################################################################################################
#######################################################################################################

#Reading library requirements
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

#OpenPyXl file Initialization
wb = load_workbook(filename)
sheet = wb.active

#Function to check what was the last line written
def findLastLineExercise(exerciseName):
    #Exercise Columns
    bpCols = [1, 2, 3, 4, 5]
    squatCols = [9,10,11,12,13]
    rowCols = [17,18,19,20,21]
    pullUpCols = [25,26,27,28,29]
    spCols = [33,34,35,36,37]
    stiffCols = [41,42,43,44,45]
    hpCols = [49,50,51,52,53]
    lungeCols = [57,58,59,60,61]

    if(exerciseName.__contains__("Bench Press") or exerciseName.__contains__("Supino")):
        chosenArr = bpCols
        chosenLine = bpLine[-1] #Last Recorded Line for Exercise
    
    elif(exerciseName.__contains__("Squat") or exerciseName.__contains__("Leg press")):
        chosenArr = squatCols
        chosenLine = squatLine[-1]

    elif(exerciseName == "Rowing (Machine)"):
        chosenArr = rowCols
        chosenLine = rowLine[-1]
        
    elif(exerciseName == "Pull Up"):
        chosenArr = pullUpCols
        chosenLine = pullUpLine[-1]

    elif(exerciseName == "Shoulder Press (Machine)"):
        chosenArr = spCols
        chosenLine = spLine[-1]

    elif(exerciseName.__contains__("Deadlift")):
        chosenArr = stiffCols   
        chosenLine = stiffLine[-1] 

    elif(exerciseName == "Hip Thrust (Barbell)"):
        chosenArr = hpCols   
        chosenLine = hpLine[-1]

    elif(exerciseName.__contains__("Lunge")):
        chosenArr = lungeCols   
        chosenLine = lungeLine[-1]

    return chosenLine

#Function to get the cellValue
def cellValue(sheet, rows, cols):
    return sheet.cell(row = rows, column = cols).value

#Function that returns the cell in which the first data from the exercise is
def findExerciseIndex(exerciseName, sheet):
    findIndexArr = []
    row_num = 2

    for col_num in range(2,60,8):
        cellExercise = str(sheet.cell(row = row_num, column = col_num).value)

        if cellExercise in exerciseName:

            serieCellIndex_col = col_num
            findIndexArr.append(serieCellIndex_col)

            serieCellIndex_row = row_num+2
            findIndexArr.append(serieCellIndex_row)

            return findIndexArr

# This function will be responsible for iterating all the series of the session and calculating volumes
def perfComparisonWritter(serieCellIndex, sheet, exerciseName):
    perfHashTable = {}

    col_s = serieCellIndex[0]
    row_s = serieCellIndex[1]

    exerciseName = exerciseNameVerifier(exerciseName)

    #Cell background color formatting variables
    cell_format = workbook.add_format()
    cell_format.set_pattern(1)  # This is optional when using a solid fill.
    cell_format.set_bg_color('orange')

    tempVolume = 0
    tempSerie  = 0
    tempWeight = 0
    tempReps   = 0

    seshVolume = 0
    seshSerie  = 0
    seshWeight = 0
    seshReps   = 0

    lastLine = findLastLineExercise(exerciseName)
    #Could iterate to insert RIGHT border until lastLine of exercise
    for ii in range(1, lastLine):
        sheet.cell(row = ii, column = col_s + 6).border = Border(right=Side(style='thick'))

    #Adjust Column Width for Relative Performance column
    RelPerfHeaders = [7,15,23,31,39,47,55,63]
    for ii in RelPerfHeaders:
        sheet.column_dimensions[get_column_letter(ii)].width = 19
    
    #Adjust Column Width for Data Column
    RelPerfHeaders = [6,14,22,30,38,46,54,62]
    for ii in RelPerfHeaders:
        sheet.column_dimensions[get_column_letter(ii)].width = 10

    i = 4 #Every first index of the exercises start in row 4
    firstWeekDate = ""

    #Iterate all lines of exerciseName columns
    while(i < lastLine + 1):

        #Session data
        if(sheet.cell(row = i, column = col_s).value != None):

            if(i == 4):
                firstWeekDate = cellValue(sheet, i, col_s + 4)

            #Add the weights x reps to volume
            tempDate = cellValue(sheet, i, col_s + 4)
            tempSerie = cellValue(sheet, i, col_s)
            tempWeight = cellValue(sheet, i, col_s + 1)
            tempReps   = cellValue(sheet, i, col_s + 2)
            tempVolume = float(tempWeight) * int(tempReps)

            #Add series to hashTable with date
            #Structure is for weights: w_SERIE_DATE
            tempWeightKey = "w_" + exerciseName + "_" + str(tempSerie) + "_" + str(tempDate)
            tempRepKey    = "r_" + exerciseName + "_" + str(tempSerie) + "_" + str(tempDate) 
            perfHashTable[tempWeightKey] = float(tempWeight)
            perfHashTable[tempRepKey] = int(tempReps)

            seshWeight += float(tempWeight)
            seshReps += int(tempReps)
            seshSerie = int(tempSerie)
            seshVolume += float(tempVolume)

            #Algorithm to add rep difference from last week session
            #ignore first week
            if(tempDate == firstWeekDate):
                #First week wont have comparisons only storage of temp values   
                print("")

            else:
                #Need to compare to PREVIOUS week only
                #Only compare if there is a equal series number of last week
                serieVerifier = "w_" + exerciseName + "_" + str(seshSerie) + "_" + str(previousWeek)

                if serieVerifier in perfHashTable.keys():

                    #Calculate difference between weeks
                    weightDiff = tempWeight - perfHashTable["w_" + exerciseName + "_" + str(seshSerie) + "_" + str(previousWeek)]
                    repDiff = tempReps - perfHashTable["r_" + exerciseName + "_" + str(seshSerie) + "_" + str(previousWeek)]

                    sheet.cell(row = i, column = col_s + 5).value = weightDiff
                    sheet.cell(row = i, column = col_s + 6).value = repDiff

                    if(weightDiff < 0):
                        #red color
                        sheet.cell(row = i, column = col_s + 5).fill = PatternFill(start_color= Color(indexed=29), fill_type = "solid")

                    elif(weightDiff == 0):
                        #yellow color
                        sheet.cell(row = i, column = col_s + 5).fill = PatternFill(start_color= Color(indexed=26), fill_type = "solid")

                    elif(weightDiff > 0):
                        #green color
                        sheet.cell(row = i, column = col_s + 5).fill = PatternFill(start_color= Color(indexed=50), fill_type = "solid")

                    if(repDiff < 0):
                        #red color
                        sheet.cell(row = i, column = col_s + 6).fill = PatternFill(start_color= Color(indexed=29), fill_type = "solid")

                    elif(repDiff == 0):
                        #yellow color
                        sheet.cell(row = i, column = col_s + 6).fill = PatternFill(start_color= Color(indexed=26), fill_type = "solid")

                    elif(repDiff > 0):
                        #green color
                        sheet.cell(row = i, column = col_s + 6).fill = PatternFill(start_color= Color(indexed=50), fill_type = "solid")

            #Increment i to pass to the next row
            i +=1

        else:
            #Calculate Medians
            weightMedian = seshWeight / seshSerie
            repsMedian = seshReps / seshSerie

            #Write volume medians
            sheet.cell(row = i, column = col_s - 1).value = "VOLUME" #Header of the row
            sheet.cell(row = i, column = col_s - 1).fill = PatternFill(start_color=Color(indexed=47), fill_type = "solid")

            sheet.cell(row = i, column = col_s).value = seshSerie
            sheet.cell(row = i, column = col_s).fill = PatternFill(start_color= Color(indexed=47), fill_type = "solid")

            sheet.cell(row = i, column = col_s+1).value = weightMedian
            sheet.cell(row = i, column = col_s+1).fill = PatternFill(start_color= Color(indexed=47), fill_type = "solid")

            sheet.cell(row = i, column = col_s+2).value = repsMedian
            sheet.cell(row = i, column = col_s+2).fill = PatternFill(start_color=Color(indexed=47), fill_type = "solid")

            sheet.cell(row = i, column = col_s+3).value = "Total Vol" 
            sheet.cell(row = i, column = col_s+3).fill = PatternFill(start_color=Color(indexed=47), fill_type = "solid") 

            sheet.cell(row = i, column = col_s+4).value = seshVolume
            sheet.cell(row = i, column = col_s+4).fill = PatternFill(start_color=Color(indexed=47), fill_type = "solid")

            #Save this week performance as previous week
            previousWeek = tempDate

            #Reset actual session data before moving to next session
            seshWeight = 0
            seshReps   = 0
            seshSerie  = 0
            seshVolume = 0
            i +=2 #Increment i to step over the next empty row

    

#INITIALIZE THE READING AND VOLUME CALCULATION AND WRITTING

#Function to choose appropriate array according to exericse Name
for exercise in exerciseList:
    cellIndex = findExerciseIndex(exercise, sheet)
    perfComparisonWritter(cellIndex, sheet, exercise)

wb.save(filename)
wb.close()
